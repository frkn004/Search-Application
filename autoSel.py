from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import platform
import random
import traceback
import os
import json
import re  # re modülünü dosya seviyesinde import ediyoruz
from datetime import datetime
import pandas as pd
import concurrent.futures
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ülkeler, diller ve eyaletler/şehirler
COUNTRIES_INFO = {
    'Türkiye': {
        'dil': 'tr',
        'eyaletler': [
            'İstanbul', 'Ankara', 'İzmir', 'Antalya', 'Adana', 'Zonguldak', 'Elazığ',
            'Konya', 'Bursa', 'Kayseri', 'Trabzon', 'Samsun', 'Erzurum', 'Gaziantep',
            'Diyarbakır', 'Kocaeli', 'Manisa', 'Mersin', 'Eskişehir', 'Şanlıurfa',
            'Malatya', 'Batman', 'Hatay', 'Afyon', 'Çanakkale', 'Denizli', 'Muğla',
            'Aydın', 'Balıkesir', 'Tekirdağ', 'Sakarya', 'Ordu', 'Rize', 'Kütahya',
            'Çorum', 'Amasya', 'Karaman', 'Kahramanmaraş', 'Erzincan', 'Tokat',
            'Isparta', 'Bolu', 'Yalova', 'Bitlis', 'Ardahan', 'Nevşehir', 'Bartın',
            'Sinop', 'Niğde', 'Mardin', 'Bingöl', 'Osmaniye', 'Kastamonu'
        ]
    },
    'Almanya': {
        'dil': 'de',
        'eyaletler': [
            'Bayern', 'Berlin', 'Hessen', 'Nordrhein-Westfalen', 'Sachsen', 
            'Baden-Württemberg', 'Hamburg', 'Niedersachsen', 'Bremen', 'Saarland', 
            'Schleswig-Holstein', 'Rheinland-Pfalz', 'Thüringen', 'Sachsen-Anhalt',
            'Brandenburg', 'Mecklenburg-Vorpommern',
            # Büyük şehirler
            'München', 'Frankfurt', 'Köln', 'Düsseldorf', 'Stuttgart', 'Dresden', 'Leipzig',
            'Hannover', 'Nürnberg', 'Dortmund', 'Essen', 'Bremen', 'Bonn', 'Wuppertal', 
            'Karlsruhe', 'Mannheim', 'Augsburg', 'Wiesbaden', 'Münster', 'Rostock'
        ]
    },
    'Fransa': {
        'dil': 'fr',
        'eyaletler': [
            'Île-de-France', 'Auvergne-Rhône-Alpes', 'Bretagne', 'Grand Est', 'Normandie', 
            'Provence-Alpes-Côte d\'Azur', 'Hauts-de-France', 'Nouvelle-Aquitaine', 
            'Occitanie', 'Pays de la Loire', 'Bourgogne-Franche-Comté', 'Centre-Val de Loire',
            'Corse',
            # Büyük şehirler
            'Paris', 'Marseille', 'Lyon', 'Toulouse', 'Nice', 'Nantes', 'Strasbourg', 
            'Montpellier', 'Bordeaux', 'Lille', 'Rennes', 'Reims', 'Le Havre', 'Saint-Étienne',
            'Toulon', 'Angers', 'Grenoble', 'Dijon', 'Le Mans', 'Clermont-Ferrand'
        ]
    },
    'İspanya': {
        'dil': 'es',
        'eyaletler': [
            'Andalucía', 'Aragón', 'Asturias', 'Cantabria', 'Castilla-La Mancha', 
            'Castilla y León', 'Cataluña', 'Comunidad Valenciana', 'Extremadura', 
            'Galicia', 'Madrid', 'Murcia', 'Navarra', 'País Vasco', 'La Rioja',
            # Büyük şehirler
            'Madrid', 'Barcelona', 'Valencia', 'Sevilla', 'Zaragoza', 'Málaga', 
            'Murcia', 'Palma de Mallorca', 'Las Palmas', 'Bilbao', 'Alicante', 
            'Córdoba', 'Valladolid', 'Vigo', 'Gijón', 'Granada', 'A Coruña', 
            'Vitoria-Gasteiz', 'Oviedo', 'Sabadell'
        ]
    },
    'İngiltere': {
        'dil': 'en',
        'eyaletler': [
            'England', 'Scotland', 'Wales', 'Northern Ireland',
            # Büyük şehirler
            'London', 'Birmingham', 'Manchester', 'Glasgow', 'Liverpool', 'Leeds',
            'Newcastle', 'Sheffield', 'Bristol', 'Edinburgh', 'Leicester', 'Coventry',
            'Cardiff', 'Belfast', 'Nottingham', 'Hull', 'Bradford', 'Aberdeen', 
            'Southampton', 'Oxford', 'Cambridge', 'York', 'Reading', 'Brighton'
        ]
    },
    'ABD': {
        'dil': 'en',
        'eyaletler': [
            'Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado', 
            'Connecticut', 'Delaware', 'Florida', 'Georgia', 'Hawaii', 'Idaho', 
            'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky', 'Louisiana', 
            'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota', 
            'Mississippi', 'Missouri', 'Montana', 'Nebraska', 'Nevada', 
            'New Hampshire', 'New Jersey', 'New Mexico', 'New York', 
            'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon', 
            'Pennsylvania', 'Rhode Island', 'South Carolina', 'South Dakota', 
            'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington', 
            'West Virginia', 'Wisconsin', 'Wyoming',
            # Büyük şehirler
            'New York City', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix', 
            'Philadelphia', 'San Antonio', 'San Diego', 'Dallas', 'San Jose', 
            'Austin', 'Jacksonville', 'San Francisco', 'Columbus', 'Charlotte',
            'Indianapolis', 'Seattle', 'Denver', 'Washington DC', 'Boston',
            'Nashville', 'Las Vegas', 'Portland', 'Detroit', 'Atlanta'
        ]
    },
    'Rusya': {
        'dil': 'ru',
        'eyaletler': [
            'Moskova', 'Sankt-Peterburg', 'Novosibirsk', 'Ekaterinburg', 'Kazan',
            'Chelyabinsk', 'Omsk', 'Samara', 'Rostov-on-Don', 'Ufa', 'Krasnoyarsk',
            'Perm', 'Voronezh', 'Volgograd', 'Krasnodar', 'Saratov', 'Tyumen',
            'Tolyatti', 'Izhevsk', 'Barnaul', 'Ulyanovsk', 'Irkutsk', 'Khabarovsk',
            'Yaroslavl', 'Vladivostok', 'Tomsk', 'Nizhny Novgorod'
        ]
    },
    'Çin': {
        'dil': 'zh',
        'eyaletler': [
            'Beijing', 'Shanghai', 'Guangzhou', 'Shenzhen', 'Tianjin', 'Chongqing',
            'Wuhan', 'Chengdu', 'Nanjing', 'Xi\'an', 'Hangzhou', 'Shenyang', 
            'Zhengzhou', 'Qingdao', 'Dalian', 'Jinan', 'Harbin', 'Changsha',
            'Suzhou', 'Ningbo', 'Wuxi', 'Fuzhou', 'Xiamen', 'Kunming', 'Dongguan',
            'Hefei', 'Changchun', 'Nanning', 'Nanchang', 'Urumqi', 'Guiyang'
        ]
    },
    'Hindistan': {
        'dil': 'hi',
        'eyaletler': [
            'Delhi', 'Mumbai', 'Kolkata', 'Chennai', 'Bangalore', 'Hyderabad',
            'Ahmedabad', 'Pune', 'Jaipur', 'Lucknow', 'Kanpur', 'Nagpur', 'Indore',
            'Thane', 'Bhopal', 'Visakhapatnam', 'Patna', 'Vadodara', 'Ludhiana',
            'Agra', 'Nashik', 'Faridabad', 'Surat', 'Coimbatore', 'Rajkot',
            'Ghaziabad', 'Varanasi', 'Amritsar', 'Chandigarh', 'Allahabad'
        ]
    },
    'Brezilya': {
        'dil': 'pt',
        'eyaletler': [
            'São Paulo', 'Rio de Janeiro', 'Brasília', 'Salvador', 'Fortaleza',
            'Belo Horizonte', 'Manaus', 'Curitiba', 'Recife', 'Porto Alegre',
            'Belém', 'Goiânia', 'Guarulhos', 'Campinas', 'São Luís', 'São Gonçalo',
            'Maceió', 'Duque de Caxias', 'Natal', 'Campo Grande', 'Teresina',
            'São Bernardo do Campo', 'Nova Iguaçu', 'João Pessoa', 'Santo André',
            'Osasco', 'São José dos Campos', 'Jaboatão dos Guararapes'
        ]
    },
    'Güney Afrika': {
        'dil': 'en',
        'eyaletler': [
            'Cape Town', 'Durban', 'Johannesburg', 'Pretoria', 'Port Elizabeth',
            'Bloemfontein', 'Nelspruit', 'Kimberley', 'Polokwane', 'Rustenburg',
            'Pietermaritzburg', 'Benoni', 'Witbank', 'Boksburg', 'Welkom',
            'Newcastle', 'Krugersdorp', 'Botshabelo', 'Richards Bay', 'Brakpan'
        ]
    },
    'Avustralya': {
        'dil': 'en',
        'eyaletler': [
            'Sydney', 'Melbourne', 'Brisbane', 'Perth', 'Adelaide', 'Gold Coast',
            'Canberra', 'Newcastle', 'Wollongong', 'Logan City', 'Geelong',
            'Hobart', 'Townsville', 'Cairns', 'Darwin', 'Toowoomba', 'Ballarat',
            'Bendigo', 'Launceston', 'Mackay', 'Rockhampton', 'Bunbury'
        ]
    },
    'Kanada': {
        'dil': 'en',
        'eyaletler': [
            'Toronto', 'Montreal', 'Vancouver', 'Calgary', 'Edmonton', 'Ottawa',
            'Quebec City', 'Winnipeg', 'Hamilton', 'Kitchener', 'London', 'Halifax',
            'Victoria', 'Windsor', 'Saskatoon', 'Regina', 'Burnaby', 'Richmond',
            'Mississauga', 'Surrey', 'Laval', 'Markham', 'Vaughan', 'Gatineau'
        ]
    },
    'Japonya': {
        'dil': 'en',  # Japonca desteklenmiyorsa İngilizce kullanılır
        'eyaletler': [
            'Tokyo', 'Yokohama', 'Osaka', 'Nagoya', 'Sapporo', 'Kobe', 'Kyoto',
            'Fukuoka', 'Kawasaki', 'Saitama', 'Hiroshima', 'Sendai', 'Kitakyushu',
            'Chiba', 'Sakai', 'Niigata', 'Hamamatsu', 'Kumamoto', 'Okayama',
            'Shizuoka', 'Kanazawa', 'Kagoshima', 'Nagasaki', 'Matsuyama'
        ]
    },
    'İtalya': {
        'dil': 'en',  # İtalyanca desteklenmiyorsa İngilizce kullanılır
        'eyaletler': [
            'Roma', 'Milano', 'Napoli', 'Torino', 'Palermo', 'Genova', 'Bologna',
            'Firenze', 'Bari', 'Catania', 'Venezia', 'Verona', 'Messina', 'Padova',
            'Trieste', 'Taranto', 'Brescia', 'Prato', 'Reggio Calabria', 'Modena',
            'Parma', 'Cagliari', 'Livorno', 'Salerno', 'Perugia', 'Rimini'
        ]
    },
    'Meksika': {
        'dil': 'es',
        'eyaletler': [
            'Mexico City', 'Guadalajara', 'Monterrey', 'Puebla', 'Tijuana', 'León',
            'Juárez', 'Zapopan', 'Ecatepec', 'Mexicali', 'Culiacán', 'Mérida',
            'Chihuahua', 'San Luis Potosí', 'Aguascalientes', 'Querétaro', 'Morelia',
            'Hermosillo', 'Saltillo', 'Cancún', 'Toluca', 'Acapulco', 'Tampico'
        ]
    },
    'Polonya': {
        'dil': 'pl',
        'eyaletler': [
            'Warszawa', 'Kraków', 'Łódź', 'Wrocław', 'Poznań', 'Gdańsk', 'Szczecin',
            'Bydgoszcz', 'Lublin', 'Katowice', 'Białystok', 'Gdynia', 'Częstochowa',
            'Radom', 'Sosnowiec', 'Toruń', 'Kielce', 'Rzeszów', 'Olsztyn', 'Gliwice'
        ]
    },
    'Güney Kore': {
        'dil': 'en',  # Korece desteklenmiyorsa İngilizce kullanılır
        'eyaletler': [
            'Seoul', 'Busan', 'Incheon', 'Daegu', 'Daejeon', 'Gwangju', 'Suwon',
            'Ulsan', 'Changwon', 'Goyang', 'Seongnam', 'Bucheon', 'Jeonju', 'Cheongju',
            'Ansan', 'Anyang', 'Cheonan', 'Hwaseong', 'Pohang', 'Jeju'
        ]
    }
}

# Farklı dillerde maden ocağı arama terimleri
MINING_TERMS = {
    'tr': ['maden ocağı', 'maden şirketi', 'maden işletmesi', 'kömür madeni', 'altın madeni'],
    'en': ['mining company', 'mine', 'mining operation', 'coal mine', 'gold mine'],
    'es': ['compañía minera', 'mina', 'operación minera', 'mina de carbón', 'mina de oro'],
    'fr': ['société minière', 'mine', 'exploitation minière', 'mine de charbon', 'mine d\'or'],
    'de': ['Bergbauunternehmen', 'Mine', 'Bergbau', 'Kohlemine', 'Goldmine'],
    'ru': ['горнодобывающая компания', 'шахта', 'горнодобывающее предприятие', 'угольная шахта', 'золотой рудник'],
    'zh': ['矿业公司', '矿山', '采矿作业', '煤矿', '金矿'],
    'sv': ['gruvföretag', 'gruva', 'gruvdrift', 'kolgruva', 'guldgruva'],
    'no': ['gruveselskap', 'gruve', 'gruvedrift', 'kullgruve', 'gullgruve'],
    'pl': ['spółka górnicza', 'kopalnia', 'działalność wydobywcza', 'kopalnia węgla', 'kopalnia złota'],
    'hi': ['खनन कंपनी', 'खदान', 'खनन संचालन', 'कोयला खदान', 'सोने की खदान'],
    'id': ['perusahaan pertambangan', 'tambang', 'operasi pertambangan', 'tambang batubara', 'tambang emas'],
    'kk': ['тау-кен компаниясы', 'шахта', 'тау-кен операциясы', 'көмір шахтасы', 'алтын кеніші'],
    'mn': ['уул уурхайн компани', 'уурхай', 'уул уурхайн үйл ажиллагаа', 'нүүрсний уурхай', 'алтны уурхай'],
    'pt': ['empresa de mineração', 'mina', 'operação de mineração', 'mina de carvão', 'mina de ouro']
}

# Sadece "kömür madeni" ve "maden ocağı" terimleri
COAL_MINE_TERMS = {
    'tr': ['kömür madeni', 'maden ocağı'],
    'en': ['coal mine', 'mining pit'],
    'es': ['mina de carbón', 'pozo minero'],
    'fr': ['mine de charbon', 'puits de mine'],
    'de': ['Kohlemine', 'Bergwerk'],
    'ru': ['угольная шахта', 'горная выработка'],
    'zh': ['煤矿', '矿井'],
    'sv': ['kolgruva', 'gruvschakt'],
    'no': ['kullgruve', 'gruveanlegg'],
    'pl': ['kopalnia węgla', 'szyb górniczy'],
    'hi': ['कोयला खदान', 'खनन गड्ढा'],
    'id': ['tambang batubara', 'lubang tambang'],
    'kk': ['көмір шахтасы', 'тау-кен өндірісі'],
    'mn': ['нүүрсний уурхай', 'уурхайн нүх'],
    'pt': ['mina de carvão', 'poço de mineração']
}

def random_sleep(min_time=1, max_time=3):
    time.sleep(random.uniform(min_time, max_time))

def search(text):
    options = Options()
    
    # Headless modunu KAPATIYORUZ - işlevsiz olduğu için kaldırıyoruz
    
    # Bot algılama sistemlerini atlatmak için ek ayarlar
    user_agent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    options.add_argument(f'user-agent={user_agent}')
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--start-maximized")
    options.add_argument("--incognito")
    options.add_argument("--disable-extensions")
    options.add_argument("--lang=tr-TR")
    options.add_argument("--disable-notifications")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("prefs", {
        "profile.default_content_setting_values.geolocation": 1,  # Konum izinlerini etkinleştir
        "profile.default_content_setting_values.notifications": 2,  # Bildirimleri devre dışı bırak
        "profile.default_content_settings.popups": 0,  # Pop-up'ları engelle
        "profile.password_manager_enabled": False  # Şifre yöneticisini devre dışı bırak
    })
    
    # İşletim sistemine göre chromedriver yolunu ayarla
    if platform.system() == 'Darwin':  # macOS
        chromedriver_path = "./chromedriver"  # Yerel dizindeki chromedriver
    else:  # Windows
        chromedriver_path = "chromedriver.exe"
    
    print(f"ChromeDriver yolu: {chromedriver_path}")
    print(f"ChromeDriver mevcut mu: {os.path.exists(chromedriver_path)}")
    
    driver = None
    data = []
    
    try:
        service = Service(executable_path=chromedriver_path)
        driver = webdriver.Chrome(service=service, options=options)
        
        # JavaScript ile webdriver özelliğini gizle
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.execute_script("Object.defineProperty(navigator, 'languages', {get: () => ['tr-TR', 'tr', 'en-US', 'en']})")
        driver.execute_script("Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] })")
        driver.execute_script("const originalQuery = window.navigator.permissions.query; window.navigator.permissions.query = (parameters) => (parameters.name === 'notifications' ? Promise.resolve({ state: Notification.permission }) : originalQuery(parameters));")
        
        print("\n=== GOOGLE MAPS ARAMA BAŞLIYOR ===")
        print(f"Arama metni: {text}")
        
        # İlk önce Google ana sayfasına git - bu bot algılamayı azaltır
        driver.get('https://www.google.com')
        random_sleep(2, 4)
        
        print("Google Maps sayfasına gidiliyor...")
        driver.get('https://www.google.com/maps')
        random_sleep(4, 6)
        
        # Tarayıcı pencere boyutunu ayarla
        driver.set_window_size(1920, 1080)
        random_sleep(1, 2)
        
        # Çerez kabul et (eğer çıkarsa)
        try:
            print("Çerez kabul penceresi kontrol ediliyor...")
            cookie_buttons = driver.find_elements(By.XPATH, "//button[contains(text(), 'Accept all') or contains(text(), 'Kabul') or contains(text(), 'Tümünü kabul et') or contains(text(), 'I agree') or contains(text(), 'Agree')]")
            if cookie_buttons:
                for button in cookie_buttons:
                    try:
                        if button.is_displayed():
                            print(f"Çerez butonu bulundu: {button.text}")
                            button.click()
                            print("Çerez kabul edildi.")
                            break
                    except Exception:
                        continue
            random_sleep(1, 2)
        except Exception as e:
            print(f"Çerez penceresi işlemi sırasında hata: {e}")
        
        print(f"Arama yapılıyor: {text}")
        try:
            # Arama kutusunun varlığını doğrula ve tıkla
            search_box_selectors = [
                "input[name='q']", 
                "#searchboxinput", 
                "input[placeholder*='Ara']", 
                "input[aria-label*='Ara']",
                "input[title*='Ara']", 
                "input[type='text']"
            ]
            search_box = None
            
            for selector in search_box_selectors:
                try:
                    print(f"Arama kutusu seçicisi deneniyor: {selector}")
                    search_box = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    if search_box and search_box.is_displayed():
                        print(f"Arama kutusu bulundu: {selector}")
                        break
                except Exception:
                    print(f"- {selector} ile bulunamadı")
                    continue
            
            if not search_box:
                print("Arama kutusu bulunamadı, manuel olarak arıyorum...")
                elements = driver.find_elements(By.TAG_NAME, "input")
                for elem in elements:
                    if elem.is_displayed() and elem.get_attribute("type") != "hidden":
                        search_box = elem
                        print(f"Manuel olarak arama kutusu bulundu: {elem.get_attribute('name') or elem.get_attribute('id') or 'isimsiz'}")
                        break
            
            if not search_box:
                raise Exception("Arama kutusu bulunamadı! Lütfen ekran görüntüsünü kontrol edin.")
            
            # Arama kutusuna tıkla ve temizle
            try:
                search_box.click()
                print("Arama kutusuna tıklandı")
                random_sleep(0.5, 1)
                
                # Bu da tıklama sorunları için alternatif bir yöntem
                driver.execute_script("arguments[0].focus();", search_box)
                random_sleep(0.5, 1)
                
                search_box.clear()
                print("Arama kutusu temizlendi")
                random_sleep(0.5, 1)
            except Exception as click_error:
                print(f"Arama kutusuna tıklama hatası: {click_error}")
                # JavaScript ile tıklama dene
                driver.execute_script("arguments[0].click();", search_box)
                driver.execute_script("arguments[0].value = '';", search_box)
            
            # İnsan benzeri yazma
            print("Arama metni giriliyor...")
            for char in text:
                search_box.send_keys(char)
                random_sleep(0.05, 0.15)
            
            print("Arama yapılıyor...")
            random_sleep(1, 2)
            search_box.send_keys(Keys.ENTER)
            print("Enter tuşuna basıldı")
            
            # Sonuçların yüklenmesi için yeterince bekle
            print("Sonuçların yüklenmesi bekleniyor...")
            random_sleep(7, 10)
            
            # Ekran görüntüsü al (hata ayıklama için)
            driver.save_screenshot("search_results.png")
            print("Ekran görüntüsü 'search_results.png' olarak kaydedildi.")
            
            # Bir süre bekle, Google'ın sonuçları yüklemesi için
            print("Sayfa tamamen yüklenene kadar ek bekleme...")
            random_sleep(5, 8)
            
            # Yeni ve güncel CSS seçicileri (2024 Mart itibariyle Google Maps'te çalışanlar)
            result_selectors = [
                "div.Nv2PK",
                "a.hfpxzc",
                "div[jsaction*='pane.resultSection'] div.bfdHYd",
                "div[jsaction*='pane.resultSection'] a",
                "div[role='feed'] div",
                "div[role='article']",
                "div.lI9IFe",
                "div.Nv2PK.THOPZb",
                "div.THOPZb",
                "div.kKVYMd",
                "div[jsaction*='mouseover:pane']"
            ]
            
            print("İşletme sonuçları aranıyor...")
            found_results = False
            
            for selector in result_selectors:
                print(f"Seçici deneniyor: {selector}")
                results = driver.find_elements(By.CSS_SELECTOR, selector)
                if results:
                    found_results = True
                    print(f"✓ {len(results)} sonuç bulundu: {selector}")
                    
                    for result in results:
                        try:
                            # Metin içeriği al
                            result_text = result.text.strip()
                            if result_text and len(result_text) > 10:  # Anlamlı içerik kontrolü
                                if result_text not in data:
                                    data.append(result_text)
                                    print(f"İşletme eklendi (yaklaşık {len(result_text)} karakter)")
                        except Exception as e:
                            print(f"Sonuç işleme hatası: {e}")
                else:
                    print(f"× Sonuç bulunamadı: {selector}")
            
            if not found_results:
                print("Hiçbir sonuç bulunamadı! Alternatif yöntem deneniyor...")
                
                # Tüm görünür metni topla
                print("Tüm görünür metinler toplanıyor...")
                visible_elements = driver.find_elements(By.XPATH, "//*[not(self::script) and not(self::style) and string-length(normalize-space(text())) > 5]")
                
                # Anlamlı metin bloklarını topla
                text_blocks = []
                current_block = []
                
                for element in visible_elements:
                    if element.is_displayed():
                        element_text = element.text.strip()
                        if element_text:
                            current_block.append(element_text)
                            # Her 3-5 metin bloğunu bir araya getir
                            if len(current_block) >= random.randint(3, 5):
                                text_blocks.append("\n".join(current_block))
                                current_block = []
                
                # Son bloğu da ekle
                if current_block:
                    text_blocks.append("\n".join(current_block))
                
                # Elde edilen blokları data listesine ekle
                for block in text_blocks:
                    if block not in data:
                        data.append(block)
                
                print(f"Alternatif yöntemle {len(text_blocks)} metin bloğu eklendi.")
            
            # Tüm yöntemler başarısız olursa en son çare olarak sayfadaki tüm metni topla
            if not data:
                print("Son çare: Sayfadaki tüm metin içeriği toplanıyor...")
                all_text = driver.find_element(By.TAG_NAME, "body").text
                # Metni boş olmayan satırlara böl
                lines = [line for line in all_text.split('\n') if line.strip()]
                
                # Her 5-10 satırı bir blok olarak birleştir
                chunks = []
                for i in range(0, len(lines), random.randint(5, 10)):
                    chunks.append('\n'.join(lines[i:i+random.randint(5, 10)]))
                
                data = chunks
                print(f"Sayfa metni {len(data)} parçaya bölündü ve eklendi.")
            
            print(f"Toplam {len(data)} veri parçası toplandı.")

        except Exception as e:
            print(f"Arama işlemi sırasında hata: {e}")
            traceback.print_exc()

    except Exception as driver_error:
        print(f"Tarayıcı başlatma hatası: {driver_error}")
        traceback.print_exc()
    
    finally:
        if driver:
            print("\nTarayıcı kapatılıyor...")
            # Ekran görüntüsü alarak bitir (hata ayıklama için)
            try:
                driver.save_screenshot("final_state.png")
                print("Son durum ekran görüntüsü 'final_state.png' olarak kaydedildi.")
            except Exception as ss_error:
                print(f"Ekran görüntüsü alma hatası: {ss_error}")
            
            random_sleep(2, 3)
            driver.quit()
    
    return data

def save_results(data, region, language):
    """Sonuçları JSON dosyasına kaydet"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"mining_companies_{region}_{language}_{timestamp}.json"
    
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Sonuçlar {filename} dosyasına kaydedildi.")
    
    return filename  # Excel dönüşümü için dosya adını döndür

def json_to_excel(json_file):
    """Tek bir JSON dosyasını Excel'e dönüştürür"""
    try:
        # JSON'ı oku
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Excel dosya adını oluştur
        excel_file = json_file.replace('.json', '.xlsx')
        
        # Veri bir liste ise, detaylı ayrıştırma yaparak daha düzenli bir DataFrame oluştur
        if isinstance(data, list):
            # Metin verilerini ayrıştır
            parsed_data = []
            
            for item in data:
                # Her metin bloğunu satırlara böl
                lines = item.split('\n')
                
                # İlk satır genellikle şirket adıdır
                company_name = lines[0] if lines else ""
                
                # Adres ve diğer bilgileri ayrıştır
                address = ""
                phone = ""
                website = ""
                rating = ""
                email = ""
                description = ""
                location_type = ""
                opening_hours = ""
                
                # Rating ve review sayısını ayır (örn: 4.5 yıldız 123 yorum)
                rating_value = ""
                review_count = ""
                
                # Telefon numarası için regex desenler
                phone_patterns = [
                    r'\+\d[\d\s\-\(\)]{5,20}',  # Uluslararası formatlar
                    r'0[\d\s\-\(\)]{5,15}',      # Türkiye formatları (0 ile başlayan)
                    r'\(\d+\)\s*\d+[\d\s\-]{5,15}',  # (555) 123 45 67 formatı
                    r'\d{3,4}[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}'  # 555 123 45 67 formatı
                ]
                
                # Web sitesi için regex deseni
                website_patterns = [
                    r'(https?://)?([a-zA-Z0-9][-a-zA-Z0-9]*\.)+[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*',
                    r'www\.[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*'
                ]
                
                # Açılış-kapanış saati için regex deseni
                hours_patterns = [
                    r'(Açık|Kapalı)\s*⋅\s*Kapanış\s*saati:\s*\d{1,2}:\d{2}',
                    r'(Açık|Kapalı)\s*⋅\s*Açılış\s*zamanı:\s*\w{2,3}\s*\d{1,2}:\d{2}',
                    r'(Açık|Kapalı|Şu anda açık|24 saat açık|Geçici olarak kapalı|Kapanmak üzere)',
                    r'\d{1,2}:\d{2}(\s*-\s*\d{1,2}:\d{2})?'
                ]
                
                # Adres için olası işaretler
                address_patterns = [
                    r'[A-Z][a-zA-ZğüşiöçĞÜŞİÖÇ]+\s+(Cad(desi)?|Sk|Sokak|Bulvarı|Mahallesi|Mah\.)',
                    r'No:\s*\d+',
                    r'Kat:?\s*\d+',
                    r'Daire:?\s*\d+'
                ]
                
                for i, line in enumerate(lines[1:], 1):
                    line = line.strip()
                    # Boş satırları atla
                    if not line:
                        continue
                    
                    line_lower = line.lower()
                    
                    # İşletme türü kontrolü
                    if i == 1 and any(type_word in line_lower for type_word in ['maden', 'ocak', 'şirket', 'mine', 'company', 'madencilik', 'müze', 'taş']):
                        if len(line) < 50:  # İşletme türü genelde kısa
                            location_type = line
                            continue
                    
                    # Puan/değerlendirme kontrolü
                    # 4,5(13) gibi formatları kontrol et - bu puan ve yorum sayısıdır, adres değil
                    if re.match(r'^\d+[,.]\d+\(\d+\)$', line):
                        import re
                        rating_match = re.search(r'(\d+[,.]\d+)', line)
                        if rating_match:
                            rating_value = rating_match.group(0)
                        
                        review_match = re.search(r'\((\d+)\)', line)
                        if review_match:
                            review_count = review_match.group(1)
                        
                        rating = f"{rating_value} puan, {review_count} yorum"
                        continue
                    
                    # Açılış-kapanış saati kontrolü
                    is_hours = False
                    for pattern in hours_patterns:
                        import re
                        if re.search(pattern, line):
                            opening_hours = line
                            is_hours = True
                            break
                    
                    if is_hours:
                        continue
                    
                    # Telefon numarası kontrolü - sadece gerçek telefon numaralarını yakala
                    is_phone = False
                    for pattern in phone_patterns:
                        import re
                        if re.search(pattern, line) and not "Kapanış saati" in line and not "Açılış zamanı" in line:
                            # Telefon numarasından önce açılış-kapanış bilgisi varsa, bu kısmı temizle
                            phone_part = re.search(pattern, line).group(0)
                            
                            # Sadece zaman içeriyor mu kontrol et (bu durumda telefon değil çalışma saati olabilir)
                            if re.match(r'^\d{1,2}:\d{2}$', phone_part):
                                continue
                                
                            phone = phone_part
                            is_phone = True
                            break
                    
                    if is_phone:
                        continue
                    
                    # E-posta kontrolü - hem doğrudan e-posta formatı hem de "e-mail: xxx@yyy.com" gibi formatları yakala
                    if '@' in line:
                        import re
                        email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', line)
                        if email_match:
                            email = email_match.group(0)
                            continue
                        
                        # E-posta etiketi varsa
                        if 'e-mail' in line_lower or 'email' in line_lower or 'e-posta' in line_lower or 'mail' in line_lower:
                            parts = line.split(':')
                            if len(parts) > 1:
                                email = parts[1].strip()
                                continue
                    
                    # Web sitesi kontrolü
                    for pattern in website_patterns:
                        import re
                        website_match = re.search(pattern, line)
                        if website_match:
                            website = website_match.group(0)
                            break
                    
                    # Eğer sadece "Web sitesi" yazıyorsa, bu bir link değil etikettir
                    if line_lower == "web sitesi" or line_lower == "website":
                        continue
                    
                    # Eğer web sitesi bulunmuşsa, sonraki adıma geç
                    if website:
                        continue
                    
                    # Adres kontrolü - tipik adres kalıpları
                    is_address = False
                    for pattern in address_patterns:
                        if re.search(pattern, line):
                            # Adres olarak işaretle ve başka bir şey olup olmadığını kontrol et
                            if not address:
                                address = line
                                is_address = True
                                break
                    
                    if is_address:
                        continue
                    
                    # Eğer adres belirlenmemişse ve bu satır işletme türünden hemen sonra geliyorsa, muhtemelen adrestir
                    if not address and i == 2 and location_type:
                        # Açılış/kapanış saati veya telefon numarası değilse adres olabilir
                        if not any(x in line_lower for x in ["açık", "kapalı"]) and not re.search(r'\d+[,.]\d+\(\d+\)', line):
                            address = line
                            continue
                    
                    # Uzun metinler açıklama olabilir
                    if len(line) > 100:
                        description += line + "\n"
                        continue
                    
                    # Kaydedilmemiş ve anlam ifade eden bir satır kaldıysa, açıklamaya ekle
                    if len(line) > 3:
                        description += line + "\n"
                
                # Ayrıştırılmış verileri ekle
                parsed_data.append({
                    'Şirket Adı': company_name,
                    'Adres': address,
                    'Telefon': phone,
                    'E-posta': email,
                    'Web Sitesi': website,
                    'Değerlendirme': rating,
                    'Açıklama': description,
                    'Tüm Bilgiler': item  # Orijinal veriyi de sakla
                })
            
            # DataFrame oluştur
            df = pd.DataFrame(parsed_data)
            
            # Sütun sıralamasını ayarla
            columns = ['Şirket Adı', 'Adres', 'Telefon', 'E-posta', 'Web Sitesi', 'Değerlendirme', 'Açıklama', 'Tüm Bilgiler']
            df = df[columns]
            
            # Excel'e yaz
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Maden Şirketleri')
                
                # Excel dosyasını güzelleştir
                workbook = writer.book
                worksheet = writer.sheets['Maden Şirketleri']
                
                # Başlık satırı stil ayarları
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                
                # Tüm hücrelere ince kenarlık ekle
                for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = border
                
                # Başlık satırını formatla
                for col_idx, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Diğer satırları formatla
                for row_idx in range(2, len(df) + 2):
                    # Satır yüksekliğini ayarla
                    worksheet.row_dimensions[row_idx].height = 18
                    
                    # Alternatif satır renklendirme
                    fill = PatternFill(start_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    end_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    fill_type="solid")
                    
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.fill = fill
                
                # Sütun genişliklerini ayarla
                for idx, col in enumerate(df.columns, 1):
                    column_letter = get_column_letter(idx)
                    if col == 'Tüm Bilgiler':
                        worksheet.column_dimensions[column_letter].width = 100
                    elif col in ['Açıklama', 'Adres', 'Arama Sorgusu']:
                        worksheet.column_dimensions[column_letter].width = 40
                    elif col in ['Şirket Adı', 'Web Sitesi']:
                        worksheet.column_dimensions[column_letter].width = 30
                    else:
                        # Diğer sütunların genişliklerini içeriğe göre ayarla (min 15, max 25)
                        column_values = df[col].astype(str)
                        max_length = max(column_values.map(len).max(), len(col)) + 3
                        worksheet.column_dimensions[column_letter].width = min(max(max_length, 15), 25)
                
                # Otomatik filtre ekle
                worksheet.auto_filter.ref = worksheet.dimensions
        
        # Veri bir sözlük ise, yapılandırılmış veriler için daha iyi bir format oluştur
        elif isinstance(data, dict):
            excel_data = []
            
            for region, countries in data.items():
                for country, languages in countries.items():
                    for lang, data_list in languages.items():
                        for item in data_list:
                            # Metin verilerini böl
                            lines = item.split('\n') if isinstance(item, str) else [""]
                            company_name = lines[0] if lines else ""
                            
                            # Adres ve diğer bilgileri ayrıştır
                            address = ""
                            phone = ""
                            website = ""
                            rating = ""
                            email = ""
                            
                            for line in lines[1:]:
                                line = line.strip()
                                # Boş satırları atla
                                if not line:
                                    continue
                                    
                                # Telefon numarası gibi gözüküyorsa
                                if any(c in line for c in ['+', '(', ')']):
                                    if len(line) < 30 and any(c.isdigit() for c in line):
                                        phone = line
                                        continue
                                
                                # Web sitesi gibi gözüküyorsa
                                if any(domain in line.lower() for domain in ['.com', '.org', '.net', 'www.', 'http']):
                                    website = line
                                    continue
                                
                                # Puan/değerlendirme gibi gözüküyorsa
                                if 'yıldız' in line.lower() or 'star' in line.lower() or 'puan' in line.lower():
                                    rating = line
                                    # Puanı ve yorum sayısını ayıkla
                                    rating_parts = line.split()
                                    for part in rating_parts:
                                        if part.replace('.', '').isdigit() or (part.count('.') == 1 and part.replace('.', '').isdigit()):
                                            try:
                                                float_val = float(part)
                                                if float_val <= 5.0:
                                                    rating_value = part
                                            except Exception:
                                                pass
                                        elif part.isdigit():
                                            review_count = part
                                    continue
                                
                                # Diğer satırlar muhtemelen adres
                                if not address and len(line) > 5:
                                    address = line
                            
                            excel_data.append({
                                "Bölge": region,
                                "Ülke": country,
                                "Dil": lang,
                                "Şirket Adı": company_name,
                                "Adres": address,
                                "Telefon": phone,
                                "E-posta": email,
                                "Web Sitesi": website,
                                "Değerlendirme": rating,
                                "Tüm Bilgiler": item
                            })
            
            if excel_data:
                df = pd.DataFrame(excel_data)
                
                # Excel'e yaz
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Maden Şirketleri')
                    
                    # Excel dosyasını güzelleştir
                    workbook = writer.book
                    worksheet = writer.sheets['Maden Şirketleri']
                    
                    # Başlık satırı stil ayarları
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True, size=12)
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    border = Border(
                        left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000')
                    )
                    
                    # Tüm hücrelere ince kenarlık ekle
                    for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                        for cell in row:
                            cell.border = border
                    
                    # Başlık satırını formatla
                    for col_idx, column in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_idx)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                    
                    # Diğer satırları formatla
                    for row_idx in range(2, len(df) + 2):
                        # Satır yüksekliğini ayarla
                        worksheet.row_dimensions[row_idx].height = 18
                        
                        # Alternatif satır renklendirme
                        fill = PatternFill(start_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                        end_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                        fill_type="solid")
                        
                        for col_idx in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                            cell.fill = fill
                    
                    # Sütun genişliklerini ayarla
                    for idx, col in enumerate(df.columns, 1):
                        if col == 'Tüm Bilgiler':
                            worksheet.column_dimensions[get_column_letter(idx)].width = 100
                        else:
                            # Diğer sütunların genişliklerini içeriğe göre ayarla (min 12, max 30)
                            column_values = df[col].astype(str)
                            max_length = max(column_values.map(len).max(), len(col)) + 3
                            worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length, 12), 30)
                    
                    # Otomatik filtre ekle
                    worksheet.auto_filter.ref = worksheet.dimensions
        
        return f"{json_file} dosyası {excel_file} olarak güzel bir şekilde dönüştürüldü."
        
    except Exception as e:
        return f"{json_file} dosyasını dönüştürürken hata: {e}"

def search_mining_companies():
    """Tüm dünyadaki maden şirketlerini arayan ana fonksiyon - şehir/eyalet bazlı arama ile"""
    all_results = {}
    json_files = []  # Excel'e dönüştürülecek dosyaların listesi
    
    for country, info in COUNTRIES_INFO.items():
        country_lang = info['dil']
        print(f"\n=== {country} ÜLKESİ ARANIYOR (Dil: {country_lang}) ===")
        country_results = {}
        
        # Her ülkenin kendi dilindeki arama terimlerini kullan
        if country_lang in MINING_TERMS:
            search_terms = MINING_TERMS[country_lang]
        else:
            # Eğer dil desteklenmiyorsa, İngilizce terimleri kullan
            search_terms = MINING_TERMS['en']
            print(f"Uyarı: {country_lang} dili desteklenmiyor, İngilizce terimler kullanılıyor.")
        
        # Önce genel ülke araması
        country_lang_results = []
        for term in search_terms:
            search_query = f"{term} {country}"
            print(f"Ülke Araması: {search_query}")
            
            try:
                results = search(search_query)
                if results:
                    country_lang_results.extend(results)
                    print(f"{len(results)} sonuç bulundu")
            except Exception as e:
                print(f"Hata oluştu: {e}")
                continue
        
        if country_lang_results:
            # Genel ülke sonuçlarını kaydet
            json_file = save_results(country_lang_results, f"{country}_genel", country_lang)
            json_files.append(json_file)
            
            # Ülkenin genel sonuçlarını ekle
            country_results["genel"] = country_lang_results
            
        # Şimdi eyalet/şehir bazlı aramalar
        state_results = {}
        for state in info['eyaletler']:
            print(f"\n--- {state} eyaleti/şehri için arama yapılıyor ---")
            state_lang_results = []
            
            for term in search_terms:
                search_query = f"{term} {state} {country}"
                print(f"Eyalet Araması: {search_query}")
                
                try:
                    results = search(search_query)
                    if results:
                        state_lang_results.extend(results)
                        print(f"{len(results)} sonuç bulundu")
                except Exception as e:
                    print(f"Hata oluştu: {e}")
                    continue
            
            if state_lang_results:
                state_results[state] = state_lang_results
                # Eyalet sonuçlarını kaydet
                json_file = save_results(state_lang_results, f"{country}_{state}", country_lang)
                json_files.append(json_file)
        
        if state_results:
            country_results["eyaletler"] = state_results
        
        if country_results:
            all_results[country] = country_results
    
    # Tüm sonuçları kaydet
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = f"all_mining_companies_{timestamp}.json"
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    print(f"\nTüm sonuçlar {json_filename} dosyasına kaydedildi.")
    json_files.append(json_filename)
    
    # JSON dosyalarını paralel olarak Excel'e dönüştür
    print("\nJSON dosyaları paralel olarak Excel'e dönüştürülüyor...")
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(json_to_excel, file): file for file in json_files}
        for future in concurrent.futures.as_completed(future_to_file):
            print(future.result())
    
    # Tek bir büyük Excel dosyası oluştur
    create_combined_excel(all_results, timestamp)
    
    return all_results

def create_combined_excel(all_results, timestamp):
    """Tüm sonuçları tek bir büyük Excel dosyasına dönüştürür"""
    try:
        print("\nTüm sonuçlar tek bir Excel dosyasına dönüştürülüyor...")
        
        # Sonuçları düzenli bir veri yapısına dönüştür
        excel_data = []
        
        for country, country_data in all_results.items():
            country_lang = COUNTRIES_INFO.get(country, {}).get('dil', 'en')
            
            # Genel ülke sonuçları
            if "genel" in country_data:
                for item in country_data["genel"]:
                    # Metin verilerini böl
                    lines = item.split('\n') if isinstance(item, str) else [""]
                    company_name = lines[0] if lines else ""
                    
                    # Adres ve diğer bilgileri ayrıştır
                    address = ""
                    phone = ""
                    website = ""
                    rating = ""
                    email = ""
                    description = ""
                    location_type = ""
                    opening_hours = ""
                    
                    # Rating ve review sayısını ayır (örn: 4.5 yıldız 123 yorum)
                    rating_value = ""
                    review_count = ""
                    
                    # Telefon numarası için regex desenler
                    phone_patterns = [
                        r'\+\d[\d\s\-\(\)]{5,20}',  # Uluslararası formatlar
                        r'0[\d\s\-\(\)]{5,15}',      # Türkiye formatları (0 ile başlayan)
                        r'\(\d+\)\s*\d+[\d\s\-]{5,15}',  # (555) 123 45 67 formatı
                        r'\d{3,4}[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}'  # 555 123 45 67 formatı
                    ]
                    
                    # Web sitesi için regex deseni
                    website_patterns = [
                        r'(https?://)?([a-zA-Z0-9][-a-zA-Z0-9]*\.)+[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*',
                        r'www\.[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*'
                    ]
                    
                    # Açılış-kapanış saati için regex deseni
                    hours_patterns = [
                        r'(Açık|Kapalı)\s*⋅\s*Kapanış\s*saati:\s*\d{1,2}:\d{2}',
                        r'(Açık|Kapalı)\s*⋅\s*Açılış\s*zamanı:\s*\w{2,3}\s*\d{1,2}:\d{2}',
                        r'(Açık|Kapalı|Şu anda açık|24 saat açık|Geçici olarak kapalı|Kapanmak üzere)',
                        r'\d{1,2}:\d{2}(\s*-\s*\d{1,2}:\d{2})?'
                    ]
                    
                    # Adres için olası işaretler
                    address_patterns = [
                        r'[A-Z][a-zA-ZğüşiöçĞÜŞİÖÇ]+\s+(Cad(desi)?|Sk|Sokak|Bulvarı|Mahallesi|Mah\.)',
                        r'No:\s*\d+',
                        r'Kat:?\s*\d+',
                        r'Daire:?\s*\d+'
                    ]
                    
                    for i, line in enumerate(lines[1:], 1):
                        line = line.strip()
                        # Boş satırları atla
                        if not line:
                            continue
                        
                        line_lower = line.lower()
                        
                        # İşletme türü kontrolü
                        if i == 1 and any(type_word in line_lower for type_word in ['maden', 'ocak', 'şirket', 'mine', 'company', 'madencilik', 'müze', 'taş']):
                            if len(line) < 50:  # İşletme türü genelde kısa
                                location_type = line
                                continue
                        
                        # Puan/değerlendirme kontrolü
                        # 4,5(13) gibi formatları kontrol et - bu puan ve yorum sayısıdır, adres değil
                        if re.match(r'^\d+[,.]\d+\(\d+\)$', line):
                            import re
                            rating_match = re.search(r'(\d+[,.]\d+)', line)
                            if rating_match:
                                rating_value = rating_match.group(0)
                            
                            review_match = re.search(r'\((\d+)\)', line)
                            if review_match:
                                review_count = review_match.group(1)
                            
                            rating = f"{rating_value} puan, {review_count} yorum"
                            continue
                        
                        # Açılış-kapanış saati kontrolü
                        is_hours = False
                        for pattern in hours_patterns:
                            import re
                            if re.search(pattern, line):
                                opening_hours = line
                                is_hours = True
                                break
                        
                        if is_hours:
                            continue
                        
                        # Telefon numarası kontrolü - sadece gerçek telefon numaralarını yakala
                        is_phone = False
                        for pattern in phone_patterns:
                            import re
                            if re.search(pattern, line) and not "Kapanış saati" in line and not "Açılış zamanı" in line:
                                # Telefon numarasından önce açılış-kapanış bilgisi varsa, bu kısmı temizle
                                phone_part = re.search(pattern, line).group(0)
                                
                                # Sadece zaman içeriyor mu kontrol et (bu durumda telefon değil çalışma saati olabilir)
                                if re.match(r'^\d{1,2}:\d{2}$', phone_part):
                                    continue
                                    
                                phone = phone_part
                                is_phone = True
                                break
                        
                        if is_phone:
                            continue
                            
                        # E-posta kontrolü - hem doğrudan e-posta formatı hem de "e-mail: xxx@yyy.com" gibi formatları yakala
                        if '@' in line:
                            import re
                            email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', line)
                            if email_match:
                                email = email_match.group(0)
                                continue
                            
                            # E-posta etiketi varsa
                            if 'e-mail' in line_lower or 'email' in line_lower or 'e-posta' in line_lower or 'mail' in line_lower:
                                parts = line.split(':')
                                if len(parts) > 1:
                                    email = parts[1].strip()
                                    continue
                        
                        # Web sitesi kontrolü
                        for pattern in website_patterns:
                            import re
                            website_match = re.search(pattern, line)
                            if website_match:
                                website = website_match.group(0)
                                break
                        
                        # Eğer sadece "Web sitesi" yazıyorsa, bu bir link değil etikettir
                        if line_lower == "web sitesi" or line_lower == "website":
                            continue
                        
                        # Eğer web sitesi bulunmuşsa, sonraki adıma geç
                        if website:
                            continue
                        
                        # Adres kontrolü - tipik adres kalıpları
                        is_address = False
                        for pattern in address_patterns:
                            if re.search(pattern, line):
                                # Adres olarak işaretle ve başka bir şey olup olmadığını kontrol et
                                if not address:
                                    address = line
                                    is_address = True
                                    break
                    
                        if is_address:
                            continue
                        
                        # Eğer adres belirlenmemişse ve bu satır işletme türünden hemen sonra geliyorsa, muhtemelen adrestir
                        if not address and i == 2 and location_type:
                            # Açılış/kapanış saati veya telefon numarası değilse adres olabilir
                            if not any(x in line_lower for x in ["açık", "kapalı"]) and not re.search(r'\d+[,.]\d+\(\d+\)', line):
                                address = line
                                continue
                        
                        # Uzun metinler açıklama olabilir
                        if len(line) > 100:
                            description += line + "\n"
                            continue
                        
                        # Kaydedilmemiş ve anlam ifade eden bir satır kaldıysa, açıklamaya ekle
                        if len(line) > 3:
                            description += line + "\n"
                        
                        # Ayrıştırılmış verileri ekle
                        excel_data.append({
                            'Ülke': country,
                            'Şirket Adı': company_name,
                            'Adres': address,
                            'Telefon': phone,
                            'E-posta': email,
                            'Web Sitesi': website,
                            'Değerlendirme': rating,
                            'Açıklama': description,
                            'Tüm Bilgiler': item
                        })
            
            # Eyalet/şehir sonuçları
            if "eyaletler" in country_data:
                for state, state_data in country_data["eyaletler"].items():
                    for item in state_data:
                        # Metin verilerini böl
                        lines = item.split('\n') if isinstance(item, str) else [""]
                        company_name = lines[0] if lines else ""
                        
                        # Adres ve diğer bilgileri ayrıştır
                        address = ""
                        phone = ""
                        website = ""
                        rating = ""
                        email = ""
                        description = ""
                        location_type = ""
                        opening_hours = ""
                        
                        # Rating ve review sayısını ayır (örn: 4.5 yıldız 123 yorum)
                        rating_value = ""
                        review_count = ""
                        
                        # Telefon numarası için regex desenler
                        phone_patterns = [
                            r'\+\d[\d\s\-\(\)]{5,20}',  # Uluslararası formatlar
                            r'0[\d\s\-\(\)]{5,15}',      # Türkiye formatları (0 ile başlayan)
                            r'\(\d+\)\s*\d+[\d\s\-]{5,15}',  # (555) 123 45 67 formatı
                            r'\d{3,4}[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}'  # 555 123 45 67 formatı
                        ]
                        
                        # Web sitesi için regex deseni
                        website_patterns = [
                            r'(https?://)?([a-zA-Z0-9][-a-zA-Z0-9]*\.)+[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*',
                            r'www\.[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*'
                        ]
                        
                        # Açılış-kapanış saati için regex deseni
                        hours_patterns = [
                            r'(Açık|Kapalı)\s*⋅\s*Kapanış\s*saati:\s*\d{1,2}:\d{2}',
                            r'(Açık|Kapalı)\s*⋅\s*Açılış\s*zamanı:\s*\w{2,3}\s*\d{1,2}:\d{2}',
                            r'(Açık|Kapalı|Şu anda açık|24 saat açık|Geçici olarak kapalı|Kapanmak üzere)',
                            r'\d{1,2}:\d{2}(\s*-\s*\d{1,2}:\d{2})?'
                        ]
                        
                        # Adres için olası işaretler
                        address_patterns = [
                            r'[A-Z][a-zA-ZğüşiöçĞÜŞİÖÇ]+\s+(Cad(desi)?|Sk|Sokak|Bulvarı|Mahallesi|Mah\.)',
                            r'No:\s*\d+',
                            r'Kat:?\s*\d+',
                            r'Daire:?\s*\d+'
                        ]
                        
                        for i, line in enumerate(lines[1:], 1):
                            line = line.strip()
                            # Boş satırları atla
                            if not line:
                                continue
                            
                            line_lower = line.lower()
                            
                            # İşletme türü kontrolü
                            if i == 1 and any(type_word in line_lower for type_word in ['maden', 'ocak', 'şirket', 'mine', 'company', 'madencilik', 'müze', 'taş']):
                                if len(line) < 50:  # İşletme türü genelde kısa
                                    location_type = line
                                    continue
                            
                            # Puan/değerlendirme kontrolü
                            # 4,5(13) gibi formatları kontrol et - bu puan ve yorum sayısıdır, adres değil
                            if re.match(r'^\d+[,.]\d+\(\d+\)$', line):
                                import re
                                rating_match = re.search(r'(\d+[,.]\d+)', line)
                                if rating_match:
                                    rating_value = rating_match.group(0)
                                
                                review_match = re.search(r'\((\d+)\)', line)
                                if review_match:
                                    review_count = review_match.group(1)
                                
                                rating = f"{rating_value} puan, {review_count} yorum"
                                continue
                            
                            # Açılış-kapanış saati kontrolü
                            is_hours = False
                            for pattern in hours_patterns:
                                import re
                                if re.search(pattern, line):
                                    opening_hours = line
                                    is_hours = True
                                    break
                            
                            if is_hours:
                                continue
                            
                            # Telefon numarası kontrolü - sadece gerçek telefon numaralarını yakala
                            is_phone = False
                            for pattern in phone_patterns:
                                import re
                                if re.search(pattern, line) and not "Kapanış saati" in line and not "Açılış zamanı" in line:
                                    # Telefon numarasından önce açılış-kapanış bilgisi varsa, bu kısmı temizle
                                    phone_part = re.search(pattern, line).group(0)
                                    
                                    # Sadece zaman içeriyor mu kontrol et (bu durumda telefon değil çalışma saati olabilir)
                                    if re.match(r'^\d{1,2}:\d{2}$', phone_part):
                                        continue
                                        
                                    phone = phone_part
                                    is_phone = True
                                    break
                            
                            if is_phone:
                                continue
                                
                            # E-posta kontrolü - hem doğrudan e-posta formatı hem de "e-mail: xxx@yyy.com" gibi formatları yakala
                            if '@' in line:
                                import re
                                email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', line)
                                if email_match:
                                    email = email_match.group(0)
                                    continue
                                
                                # E-posta etiketi varsa
                                if 'e-mail' in line_lower or 'email' in line_lower or 'e-posta' in line_lower or 'mail' in line_lower:
                                    parts = line.split(':')
                                    if len(parts) > 1:
                                        email = parts[1].strip()
                                        continue
                            
                            # Web sitesi kontrolü
                            for pattern in website_patterns:
                                import re
                                website_match = re.search(pattern, line)
                                if website_match:
                                    website = website_match.group(0)
                                    break
                            
                            # Eğer sadece "Web sitesi" yazıyorsa, bu bir link değil etikettir
                            if line_lower == "web sitesi" or line_lower == "website":
                                continue
                            
                            # Eğer web sitesi bulunmuşsa, sonraki adıma geç
                            if website:
                                continue
                            
                            # Adres kontrolü - tipik adres kalıpları
                            is_address = False
                            for pattern in address_patterns:
                                if re.search(pattern, line):
                                    # Adres olarak işaretle ve başka bir şey olup olmadığını kontrol et
                                    if not address:
                                        address = line
                                        is_address = True
                                        break
                            
                            if is_address:
                                continue
                            
                            # Eğer adres belirlenmemişse ve bu satır işletme türünden hemen sonra geliyorsa, muhtemelen adrestir
                            if not address and i == 2 and location_type:
                                # Açılış/kapanış saati veya telefon numarası değilse adres olabilir
                                if not any(x in line_lower for x in ["açık", "kapalı"]) and not re.search(r'\d+[,.]\d+\(\d+\)', line):
                                    address = line
                                    continue
                            
                            # Uzun metinler açıklama olabilir
                            if len(line) > 100:
                                description += line + "\n"
                                continue
                            
                            # Kaydedilmemiş ve anlam ifade eden bir satır kaldıysa, açıklamaya ekle
                            if len(line) > 3:
                                description += line + "\n"
                        
                        # Ayrıştırılmış verileri ekle
                        excel_data.append({
                            'Ülke': country,
                            'Eyalet/Şehir': state,
                            'Şirket Adı': company_name,
                            'Adres': address,
                            'Telefon': phone,
                            'E-posta': email,
                            'Web Sitesi': website,
                            'Değerlendirme': rating,
                            'Açıklama': description,
                            'Tüm Bilgiler': item
                        })
        
        if excel_data:
            # DataFrame'e dönüştür
            df = pd.DataFrame(excel_data)
            
            # Excel dosyasını oluştur
            excel_filename = f"all_mining_companies_combined_{timestamp}.xlsx"
            
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Tüm Maden Şirketleri')
                
                # Excel dosyasını güzelleştir
                workbook = writer.book
                worksheet = writer.sheets['Tüm Maden Şirketleri']
                
                # Başlık satırı stil ayarları
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                
                # Tüm hücrelere ince kenarlık ekle
                for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = border
                
                # Başlık satırını formatla
                for col_idx, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Diğer satırları formatla
                for row_idx in range(2, len(df) + 2):
                    # Satır yüksekliğini ayarla
                    worksheet.row_dimensions[row_idx].height = 18
                    
                    # Alternatif satır renklendirme
                    fill = PatternFill(start_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    end_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    fill_type="solid")
                    
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.fill = fill
                
                # Sütun genişliklerini ayarla
                for idx, col in enumerate(df.columns, 1):
                    column_letter = get_column_letter(idx)
                    if col == 'Tüm Bilgiler':
                        worksheet.column_dimensions[column_letter].width = 100
                    elif col in ['Açıklama', 'Adres', 'Arama Sorgusu']:
                        worksheet.column_dimensions[column_letter].width = 40
                    elif col in ['Şirket Adı', 'Web Sitesi']:
                        worksheet.column_dimensions[column_letter].width = 30
                    else:
                        # Diğer sütunların genişliklerini içeriğe göre ayarla (min 15, max 25)
                        column_values = df[col].astype(str)
                        max_length = max(column_values.map(len).max(), len(col)) + 3
                        worksheet.column_dimensions[column_letter].width = min(max(max_length, 15), 25)
                
                # Otomatik filtre ekle
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Dondurulmuş başlık satırı
                worksheet.freeze_panes = 'A2'
            
            print(f"Tüm sonuçlar {excel_filename} dosyasına dönüştürüldü.")
        else:
            print("Excel'e dönüştürülecek veri bulunamadı.")

    except Exception as e:
        print(f"Birleştirilmiş Excel oluşturma hatası: {e}")
        traceback.print_exc()

def convert_json_to_excel():
    """Mevcut JSON dosyalarını Excel'e dönüştürür"""
    print("\nMevcut JSON dosyaları Excel'e dönüştürülüyor...")
    
    # mining_companies_ ile başlayan tüm JSON dosyalarını bul
    json_files = [f for f in os.listdir('.') if (f.startswith('mining_companies_') or f.startswith('all_mining_companies_')) and f.endswith('.json')]
    
    if not json_files:
        print("Dönüştürülecek JSON dosyası bulunamadı.")
        return
    
    print(f"{len(json_files)} JSON dosyası bulundu. Paralel olarak dönüştürülüyor...")
    
    # Paralel olarak dönüştür
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(json_to_excel, file): file for file in json_files}
        for future in concurrent.futures.as_completed(future_to_file):
            print(future.result())
    
    print("Dönüştürme işlemi tamamlandı.")

def search_one_city():
    """Tek bir şehir için arama yapar"""
    # global re modülünü kullanalım
    global re

    # COUNTRIES_INFO sözlüğünden Türkiye ve Zonguldak şehrini seç
    country = 'Türkiye'
    city = 'Zonguldak'
    country_info = COUNTRIES_INFO[country]
    country_lang = country_info['dil']
    
    print(f"\n=== TEK ŞEHİR İÇİN MADEN ŞİRKETLERİ ARAMASI ===\n")
    print(f"=== {country} ülkesinde {city} şehri için arama yapılıyor (Dil: {country_lang}) ===")
    
    search_terms = MINING_TERMS[country_lang]
    
    # Her arama terimi için tek bir sorgu oluştur
    all_results = []
    all_search_queries = []
    
    # Her arama terimi için şehir bazlı arama yap
    for term in search_terms:
        try:
            search_query = f"{term} {city} {country}"
            all_search_queries.append(search_query)
            
            print(f"Arama: {search_query}")
            
            # Aramayı yap ve sonuçları al
            results = search(search_query)
            
            # Tekrar eden sonuçları filtrele
            unique_results = []
            
            for result in results:
                # Şirket adını al (genellikle ilk satır)
                company_name = result.split('\n')[0].strip() if result else ""
                
                # Tekrar ediyorsa atla
                if company_name and any(company_name == existing.split('\n')[0].strip() for existing in unique_results):
                    print(f"Tekrar eden şirket atlandı: {company_name}")
                    continue
                
                unique_results.append(result)
                
                # Sonuçları ekle
                all_results.append({
                    'query': search_query,
                    'term': term,
                    'result': result
                })
            print(f"{len(results)} sonuç bulundu (tekrar edenler filtrelendi)")
        except Exception as e:
            print(f"Hata oluştu: {e}")
            continue
    
    # Sonuçları tek bir Excel dosyasına kaydet
    if all_results:
        print(f"\nToplam {len(all_results)} benzersiz sonuç bulundu.")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Çalışma dizinini doğrula
        working_dir = os.getcwd()
        print(f"Çalışma dizini: {working_dir}")
        
        # Önce JSON'a kaydet (backup için)
        json_filename = f"mining_companies_{country}_{city}_{timestamp}.json"
        json_path = os.path.join(working_dir, json_filename)
        
        with open(json_path, 'w', encoding='utf-8') as f:
            # JSON verilerini düzenli formatta kaydet
            json_data = {
                'ülke': country,
                'şehir': city,
                'dil': country_lang,
                'arama_terimleri': search_terms,
                'arama_sorguları': all_search_queries,
                'sonuçlar': all_results
            }
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        print(f"Sonuçlar {json_path} dosyasına kaydedildi.")
        
        # Sonuçları Excel'e dönüştür
        excel_filename = f"mining_companies_{country}_{city}_{timestamp}.xlsx"
        excel_path = os.path.join(working_dir, excel_filename)
        
        # Metin verilerini ayrıştır
        parsed_data = []
        
        for item in all_results:
            search_query = item['query']
            search_term = item['term']
            result_text = item['result']
            
            # Her metin bloğunu satırlara böl
            lines = result_text.split('\n')
            
            # İlk satır genellikle şirket adıdır
            company_name = lines[0] if lines else ""
            
            # Adres ve diğer bilgileri ayrıştır
            address = ""
            phone = ""
            website = ""
            rating = ""
            email = ""
            description = ""
            location_type = ""
            opening_hours = ""
            
            # Rating ve review sayısını ayır (örn: 4.5 yıldız 123 yorum)
            rating_value = ""
            review_count = ""
            
            # Telefon numarası için regex desenler
            phone_patterns = [
                r'\+\d[\d\s\-\(\)]{5,20}',  # Uluslararası formatlar
                r'0[\d\s\-\(\)]{5,15}',      # Türkiye formatları (0 ile başlayan)
                r'\(\d+\)\s*\d+[\d\s\-]{5,15}',  # (555) 123 45 67 formatı
                r'\d{3,4}[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}'  # 555 123 45 67 formatı
            ]
            
            # Web sitesi için regex deseni
            website_patterns = [
                r'(https?://)?([a-zA-Z0-9][-a-zA-Z0-9]*\.)+[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*',
                r'www\.[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*'
            ]
            
            # Açılış-kapanış saati için regex deseni
            hours_patterns = [
                r'(Açık|Kapalı)\s*⋅\s*Kapanış\s*saati:\s*\d{1,2}:\d{2}',
                r'(Açık|Kapalı)\s*⋅\s*Açılış\s*zamanı:\s*\w{2,3}\s*\d{1,2}:\d{2}',
                r'(Açık|Kapalı|Şu anda açık|24 saat açık|Geçici olarak kapalı|Kapanmak üzere)',
                r'\d{1,2}:\d{2}(\s*-\s*\d{1,2}:\d{2})?'
            ]
            
            # Adres için olası işaretler
            address_patterns = [
                r'[A-Z][a-zA-ZğüşiöçĞÜŞİÖÇ]+\s+(Cad(desi)?|Sk|Sokak|Bulvarı|Mahallesi|Mah\.)',
                r'No:\s*\d+',
                r'Kat:?\s*\d+',
                r'Daire:?\s*\d+'
            ]
            
            for i, line in enumerate(lines[1:], 1):
                line = line.strip()
                # Boş satırları atla
                if not line:
                    continue
                
                line_lower = line.lower()
                
                # İşletme türü kontrolü
                if i == 1 and any(type_word in line_lower for type_word in ['maden', 'ocak', 'şirket', 'mine', 'company', 'madencilik', 'müze', 'taş']):
                    if len(line) < 50:  # İşletme türü genelde kısa
                        location_type = line
                        continue
                
                # Puan/değerlendirme kontrolü
                # 4,5(13) gibi formatları kontrol et - bu puan ve yorum sayısıdır, adres değil
                if re.match(r'^\d+[,.]\d+\(\d+\)$', line):
                    rating_match = re.search(r'(\d+[,.]\d+)', line)
                    if rating_match:
                        rating_value = rating_match.group(0)
                    
                    review_match = re.search(r'\((\d+)\)', line)
                    if review_match:
                        review_count = review_match.group(1)
                    
                    rating = f"{rating_value} puan, {review_count} yorum"
                    continue
                
                # Açılış-kapanış saati kontrolü
                is_hours = False
                for pattern in hours_patterns:
                    import re
                    if re.search(pattern, line):
                        opening_hours = line
                        is_hours = True
                        break
                
                if is_hours:
                    continue
                
                # Telefon numarası kontrolü - sadece gerçek telefon numaralarını yakala
                is_phone = False
                for pattern in phone_patterns:
                    import re
                    if re.search(pattern, line) and not "Kapanış saati" in line and not "Açılış zamanı" in line:
                        # Telefon numarasından önce açılış-kapanış bilgisi varsa, bu kısmı temizle
                        phone_part = re.search(pattern, line).group(0)
                        
                        # Sadece zaman içeriyor mu kontrol et (bu durumda telefon değil çalışma saati olabilir)
                        if re.match(r'^\d{1,2}:\d{2}$', phone_part):
                            continue
                            
                        phone = phone_part
                        is_phone = True
                        break
                
                if is_phone:
                    continue
                    
                # E-posta kontrolü - hem doğrudan e-posta formatı hem de "e-mail: xxx@yyy.com" gibi formatları yakala
                if '@' in line:
                    import re
                    email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', line)
                    if email_match:
                        email = email_match.group(0)
                        continue
                    
                    # E-posta etiketi varsa
                    if 'e-mail' in line_lower or 'email' in line_lower or 'e-posta' in line_lower or 'mail' in line_lower:
                        parts = line.split(':')
                        if len(parts) > 1:
                            email = parts[1].strip()
                            continue
                
                # Web sitesi kontrolü
                for pattern in website_patterns:
                    import re
                    website_match = re.search(pattern, line)
                    if website_match:
                        website = website_match.group(0)
                        break
                
                # Eğer sadece "Web sitesi" yazıyorsa, bu bir link değil etikettir
                if line_lower == "web sitesi" or line_lower == "website":
                    continue
                
                # Eğer web sitesi bulunmuşsa, sonraki adıma geç
                if website:
                    continue
                
                # Adres kontrolü - tipik adres kalıpları
                is_address = False
                for pattern in address_patterns:
                    if re.search(pattern, line):
                        # Adres olarak işaretle ve başka bir şey olup olmadığını kontrol et
                        if not address:
                            address = line
                            is_address = True
                            break
                
                if is_address:
                    continue
                
                # Eğer adres belirlenmemişse ve bu satır işletme türünden hemen sonra geliyorsa, muhtemelen adrestir
                if not address and i == 2 and location_type:
                    # Açılış/kapanış saati veya telefon numarası değilse adres olabilir
                    if not any(x in line_lower for x in ["açık", "kapalı"]) and not re.search(r'\d+[,.]\d+\(\d+\)', line):
                        address = line
                        continue
                
                # Uzun metinler açıklama olabilir
                if len(line) > 100:
                    description += line + "\n"
                    continue
                
                # Kaydedilmemiş ve anlam ifade eden bir satır kaldıysa, açıklamaya ekle
                if len(line) > 3:
                    description += line + "\n"
            
            # Tüm Bilgiler içinde telefon ve e-posta ara (eğer hâlâ bulunamadıysa)
            if not phone:
                for pattern in phone_patterns:
                    import re
                    matches = re.findall(pattern, result_text)
                    if matches:
                        for match in matches:
                            # Açılış-kapanış saati değilse
                            if not any(x in match for x in [":", "⋅"]):
                                phone = match
                                break
                        if phone:
                            break
            
            # Telefonu tekrar kontrol et - Açılış saati yerine telefon göstermediğinden emin ol
            if phone and (re.match(r'^\d{1,2}:\d{2}$', phone) or "Kapanış saati" in phone or "Açılış zamanı" in phone):
                phone = ""
            
            # E-posta aramaya devam et
            if not email:
                import re
                email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', result_text)
                if email_matches:
                    email = email_matches[0]
            
            # Web sitesi bulunamadıysa, ilgili metni ara
            if not website:
                for pattern in website_patterns:
                    website_matches = re.findall(pattern, result_text)
                    if website_matches:
                        website = website_matches[0]
                        break
            
            # Ayrıştırılmış verileri ekle
            parsed_data.append({
                'Arama Sorgusu': search_query,
                'Arama Terimi': search_term,
                'Ülke': country,
                'Şehir': city,
                'Şirket Adı': company_name,
                'İşletme Türü': location_type,
                'Adres': address,
                'Telefon': phone,
                'E-posta': email,
                'Web Sitesi': website,
                'Değerlendirme': rating,
                'Puan': rating_value,
                'Yorum Sayısı': review_count,
                'Açıklama': description,
                'Tüm Bilgiler': result_text
            })
        
        # DataFrame oluştur
        df = pd.DataFrame(parsed_data)
        
        # Sütun sıralamasını ayarla
        columns = [
            'Ülke', 'Şehir', 'Şirket Adı', 'İşletme Türü', 'Adres', 'Telefon', 
            'E-posta', 'Web Sitesi', 'Değerlendirme', 'Puan', 'Yorum Sayısı', 
            'Açıklama', 'Arama Sorgusu', 'Arama Terimi', 'Tüm Bilgiler'
        ]
        df = df[[col for col in columns if col in df.columns]]
        
        # Excel'e yaz
        try:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Ana sonuçlar sayfası
                df.to_excel(writer, index=False, sheet_name=f'Maden Şirketleri')
                
                # Arama sorguları sayfası
                queries_df = pd.DataFrame({
                    'Arama Sorguları': all_search_queries
                })
                queries_df.to_excel(writer, index=False, sheet_name='Arama Sorguları')
                
                # Excel dosyasını güzelleştir
                workbook = writer.book
                
                # Ana sayfa formatını ayarla
                worksheet = writer.sheets['Maden Şirketleri']
                
                # Başlık satırı stil ayarları
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Kenarlık ayarları
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                
                # Tüm hücrelere ince kenarlık ekle
                for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                    for cell in row:
                        cell.border = border
                
                # Başlık satırını formatla
                for col_idx, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Diğer satırları formatla
                for row_idx in range(2, len(df) + 2):
                    # Satır yüksekliğini ayarla
                    worksheet.row_dimensions[row_idx].height = 22
                    
                    # Alternatif satır renklendirme
                    fill = PatternFill(start_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    end_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    fill_type="solid")
                    
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.fill = fill
                
                # Sütun genişliklerini ayarla
                for idx, col in enumerate(df.columns, 1):
                    column_letter = get_column_letter(idx)
                    if col == 'Tüm Bilgiler':
                        worksheet.column_dimensions[column_letter].width = 100
                    elif col in ['Açıklama', 'Adres', 'Arama Sorgusu']:
                        worksheet.column_dimensions[column_letter].width = 40
                    elif col in ['Şirket Adı', 'Web Sitesi']:
                        worksheet.column_dimensions[column_letter].width = 30
                    else:
                        # Diğer sütunların genişliklerini içeriğe göre ayarla (min 15, max 25)
                        column_values = df[col].astype(str)
                        max_length = max(column_values.map(len).max(), len(col)) + 3
                        worksheet.column_dimensions[column_letter].width = min(max(max_length, 15), 25)
                
                # Otomatik filtre ekle
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Dondurulmuş başlık satırı
                worksheet.freeze_panes = 'A2'
                
                # Arama Sorguları sayfasını formatla
                query_sheet = writer.sheets['Arama Sorguları']
                
                # Başlık satırını formatla
                cell = query_sheet.cell(row=1, column=1)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
                # Diğer satırlar
                for row_idx in range(2, len(all_search_queries) + 2):
                    fill = PatternFill(start_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    end_color="ECF0F1" if row_idx % 2 == 0 else "FFFFFF", 
                                    fill_type="solid")
                    cell = query_sheet.cell(row=row_idx, column=1)
                    cell.alignment = Alignment(vertical='center')
                    cell.fill = fill
                    cell.border = border
                
                query_sheet.column_dimensions['A'].width = 50
                
            print(f"Sonuçlar {excel_path} Excel dosyasına kaydedildi.")
            
            # Dosya varlığını doğrula
            if os.path.exists(excel_path):
                print(f"Excel dosyası başarıyla oluşturuldu: {excel_path}")
                # Dosyayı aç (Windows)
                if platform.system() == 'Windows':
                    os.startfile(excel_path)
                # Dosyayı aç (macOS)
                elif platform.system() == 'Darwin':
                    os.system(f'open "{excel_path}"')
                # Dosyayı aç (Linux)
                elif platform.system() == 'Linux':
                    os.system(f'xdg-open "{excel_path}"')
            else:
                print(f"UYARI: Excel dosyası oluşturulamadı!")
                
        except Exception as excel_error:
            print(f"Excel dosyası oluşturma hatası: {excel_error}")
            traceback.print_exc()
            
    else:
        print("Sonuç bulunamadı.")
    
    return all_results

def search_coal_mines_worldwide():
    """Dünyadaki tüm kömür madenleri ve maden ocaklarını arayan özel fonksiyon"""
    all_results = {}
    json_files = []  # Excel'e dönüştürülecek dosyaların listesi
    
    print("\n=== TÜM DÜNYADAKI KÖMÜR MADENLERİ VE MADEN OCAKLARI ARANIYOR ===\n")
    
    for country, info in COUNTRIES_INFO.items():
        country_lang = info['dil']
        print(f"\n=== {country} ÜLKESİ ARANIYOR (Dil: {country_lang}) ===")
        country_results = {}
        
        # Her ülkenin kendi dilindeki kömür madeni ve maden ocağı terimlerini kullan
        if country_lang in COAL_MINE_TERMS:
            search_terms = COAL_MINE_TERMS[country_lang]
        else:
            # Eğer dil desteklenmiyorsa, İngilizce terimleri kullan
            search_terms = COAL_MINE_TERMS['en']
            print(f"Uyarı: {country_lang} dili desteklenmiyor, İngilizce terimler kullanılıyor.")
        
        # Önce genel ülke araması
        country_lang_results = []
        for term in search_terms:
            search_query = f"{term} {country}"
            print(f"Ülke Araması: {search_query}")
            
            try:
                results = search(search_query)
                if results:
                    country_lang_results.extend(results)
                    print(f"{len(results)} sonuç bulundu")
            except Exception as e:
                print(f"Hata oluştu: {e}")
                continue
        
        if country_lang_results:
            # Genel ülke sonuçlarını kaydet
            json_file = save_results(country_lang_results, f"{country}_genel", country_lang)
            json_files.append(json_file)
            
            # Ülkenin genel sonuçlarını ekle
            country_results["genel"] = country_lang_results
            
        # Şimdi eyalet/şehir bazlı aramalar
        state_results = {}
        for state in info['eyaletler']:
            print(f"\n--- {state} eyaleti/şehri için arama yapılıyor ---")
            state_lang_results = []
            
            for term in search_terms:
                search_query = f"{term} {state} {country}"
                print(f"Eyalet Araması: {search_query}")
                
                try:
                    results = search(search_query)
                    if results:
                        state_lang_results.extend(results)
                        print(f"{len(results)} sonuç bulundu")
                except Exception as e:
                    print(f"Hata oluştu: {e}")
                    continue
            
            if state_lang_results:
                state_results[state] = state_lang_results
                # Eyalet sonuçlarını kaydet
                json_file = save_results(state_lang_results, f"{country}_{state}", country_lang)
                json_files.append(json_file)
        
        if state_results:
            country_results["eyaletler"] = state_results
        
        if country_results:
            all_results[country] = country_results
    
    # Tüm sonuçları kaydet
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_filename = f"all_coal_mines_{timestamp}.json"
    with open(json_filename, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    print(f"\nTüm sonuçlar {json_filename} dosyasına kaydedildi.")
    json_files.append(json_filename)
    
    # JSON dosyalarını paralel olarak ayrı ayrı Excel'e dönüştür
    print("\nJSON dosyaları paralel olarak Excel'e dönüştürülüyor...")
    excel_files = []
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_to_file = {executor.submit(json_to_excel, file): file for file in json_files}
        for future in concurrent.futures.as_completed(future_to_file):
            result = future.result()
            if result:
                excel_files.append(result)
                print(f"Dönüştürüldü: {result}")
    
    # Tek bir büyük Excel dosyası da oluştur
    print("\nAyrıca tüm kömür madeni sonuçları tek bir Excel dosyasına dönüştürülüyor...")
    excel_filename = create_combined_excel_coal_mines(all_results, timestamp)
    if excel_filename:
        excel_files.append(excel_filename)
        print(f"Tüm sonuçlar tek bir dosyada toplandı: {excel_filename}")
    
    # Oluşturulan tüm Excel dosyalarını listele
    print("\nOluşturulan Excel dosyaları:")
    for file in excel_files:
        print(f"- {file}")
    
    return all_results

def create_combined_excel_coal_mines(all_results, timestamp):
    """Tüm kömür madeni sonuçlarını tek bir büyük Excel dosyasına dönüştürür"""
    try:
        print("\nTüm kömür madeni sonuçları tek bir Excel dosyasına dönüştürülüyor...")
        
        # Sonuçları düzenli bir veri yapısına dönüştür
        excel_data = []
        
        for country, country_data in all_results.items():
            country_lang = COUNTRIES_INFO.get(country, {}).get('dil', 'en')
            
            # Genel ülke sonuçları
            if "genel" in country_data:
                for item in country_data["genel"]:
                    # Veri ayrıştırmayı mevcut yapılara benzer şekilde yap
                    processed_data = parse_business_data(item, country, "Genel", country_lang)
                    excel_data.append(processed_data)
            
            # Eyalet/şehir sonuçları
            if "eyaletler" in country_data:
                for state, state_data in country_data["eyaletler"].items():
                    for item in state_data:
                        # Veri ayrıştırmayı mevcut yapılara benzer şekilde yap
                        processed_data = parse_business_data(item, country, state, country_lang)
                        excel_data.append(processed_data)
        
        if excel_data:
            # DataFrame'e dönüştür
            df = pd.DataFrame(excel_data)
            
            # Excel dosyasını oluştur
            excel_filename = f"all_coal_mines_combined_{timestamp}.xlsx"
            
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Kömür Madenleri')
                
                # Excel dosyasını güzelleştir
                workbook = writer.book
                worksheet = writer.sheets['Kömür Madenleri']
                
                # Başlık satırı stil ayarları
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                
                # Sütun genişliklerini ayarla
                column_widths = {
                    'A': 15,  # Ülke
                    'B': 20,  # Eyalet/Şehir
                    'C': 10,  # Dil
                    'D': 35,  # Şirket Adı
                    'E': 20,  # İşletme Türü
                    'F': 40,  # Adres
                    'G': 20,  # Telefon
                    'H': 30,  # E-posta
                    'I': 30,  # Web Sitesi
                    'J': 20,  # Değerlendirme
                    'K': 10,  # Puan
                    'L': 15,  # Yorum Sayısı
                    'M': 50,  # Açıklama
                    'N': 20,  # Çalışma Saatleri
                    'O': 100,  # Tüm Bilgiler
                }
                
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width
                
                # Başlık satırını biçimlendir
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Veri satırlarını biçimlendir
                data_alignment = Alignment(vertical='top', wrap_text=True)
                data_border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                
                # Alternatif satır renklendirmesi
                even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(excel_data) + 1), 2):
                    # Alternatif satır renklendirmesi
                    if row_idx % 2 == 0:
                        for cell in row:
                            cell.fill = even_fill
                    
                    # Tüm hücrelere kenarlık ve hizalama ekle
                    for cell in row:
                        cell.alignment = data_alignment
                        cell.border = data_border
            
            print(f"Tüm kömür madeni sonuçları {excel_filename} dosyasına dönüştürüldü.")
            return excel_filename
        else:
            print("Excel'e dönüştürülecek veri bulunamadı.")
            return None

    except Exception as e:
        print(f"Birleştirilmiş Excel oluşturma hatası: {e}")
        traceback.print_exc()
        return None

def parse_business_data(item, country, state, language):
    """Maden şirketi verilerini ayrıştırır ve düzenli bir yapıya dönüştürür"""
    # Veri ayrıştırmayı mevcut yapılara benzer şekilde yap
    lines = item.split('\n') if isinstance(item, str) else [""]
    company_name = lines[0] if lines else ""
    
    # Adres ve diğer bilgileri ayrıştır
    address = ""
    phone = ""
    website = ""
    rating = ""
    email = ""
    description = ""
    location_type = ""
    opening_hours = ""
    
    # Rating ve review sayısını ayır (örn: 4.5 yıldız 123 yorum)
    rating_value = ""
    review_count = ""
    
    # Telefon numarası için regex desenler
    phone_patterns = [
        r'\+\d[\d\s\-\(\)]{5,20}',  # Uluslararası formatlar
        r'0[\d\s\-\(\)]{5,15}',      # Türkiye formatları (0 ile başlayan)
        r'\(\d+\)\s*\d+[\d\s\-]{5,15}',  # (555) 123 45 67 formatı
        r'\d{3,4}[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}'  # 555 123 45 67 formatı
    ]
    
    # Web sitesi için regex deseni
    website_patterns = [
        r'(https?://)?([a-zA-Z0-9][-a-zA-Z0-9]*\.)+[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*',
        r'www\.[a-zA-Z0-9][-a-zA-Z0-9]*\.[a-zA-Z]{2,}(/[-a-zA-Z0-9%_.~#+]*)*'
    ]
    
    # Açılış-kapanış saati için regex deseni
    hours_patterns = [
        r'(Açık|Kapalı)\s*⋅\s*Kapanış\s*saati:\s*\d{1,2}:\d{2}',
        r'(Açık|Kapalı)\s*⋅\s*Açılış\s*zamanı:\s*\w{2,3}\s*\d{1,2}:\d{2}',
        r'(Açık|Kapalı|Şu anda açık|24 saat açık|Geçici olarak kapalı|Kapanmak üzere)',
        r'\d{1,2}:\d{2}(\s*-\s*\d{1,2}:\d{2})?'
    ]
    
    # Adres için olası işaretler
    address_patterns = [
        r'[A-Z][a-zA-ZğüşiöçĞÜŞİÖÇ]+\s+(Cad(desi)?|Sk|Sokak|Bulvarı|Mahallesi|Mah\.)',
        r'No:\s*\d+',
        r'Kat:?\s*\d+',
        r'Daire:?\s*\d+'
    ]
    
    for i, line in enumerate(lines[1:], 1):
        line = line.strip()
        # Boş satırları atla
        if not line:
            continue
        
        line_lower = line.lower()
        
        # İşletme türü kontrolü
        if i == 1 and any(type_word in line_lower for type_word in ['maden', 'ocak', 'şirket', 'mine', 'company', 'madencilik', 'müze', 'taş', 'kömür', 'coal']):
            if len(line) < 50:  # İşletme türü genelde kısa
                location_type = line
                continue
        
        # Puan/değerlendirme kontrolü
        # 4,5(13) gibi formatları kontrol et - bu puan ve yorum sayısıdır, adres değil
        if re.match(r'^\d+[,.]\d+\(\d+\)$', line):
            rating_match = re.search(r'(\d+[,.]\d+)', line)
            if rating_match:
                rating_value = rating_match.group(0)
            
            review_match = re.search(r'\((\d+)\)', line)
            if review_match:
                review_count = review_match.group(1)
            
            rating = f"{rating_value} puan, {review_count} yorum"
            continue
        
        # Açılış-kapanış saati kontrolü
        is_hours = False
        for pattern in hours_patterns:
            if re.search(pattern, line):
                opening_hours = line
                is_hours = True
                break
        
        if is_hours:
            continue
        
        # Telefon numarası kontrolü - sadece gerçek telefon numaralarını yakala
        is_phone = False
        for pattern in phone_patterns:
            if re.search(pattern, line) and not "Kapanış saati" in line and not "Açılış zamanı" in line:
                # Telefon numarasından önce açılış-kapanış bilgisi varsa, bu kısmı temizle
                phone_part = re.search(pattern, line).group(0)
                
                # Sadece zaman içeriyor mu kontrol et (bu durumda telefon değil çalışma saati olabilir)
                if re.match(r'^\d{1,2}:\d{2}$', phone_part):
                    continue
                    
                phone = phone_part
                is_phone = True
                break
        
        if is_phone:
            continue
            
        # E-posta kontrolü - hem doğrudan e-posta formatı hem de "e-mail: xxx@yyy.com" gibi formatları yakala
        if '@' in line:
            email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', line)
            if email_match:
                email = email_match.group(0)
                continue
            
            # E-posta etiketi varsa
            if 'e-mail' in line_lower or 'email' in line_lower or 'e-posta' in line_lower or 'mail' in line_lower:
                parts = line.split(':')
                if len(parts) > 1:
                    email = parts[1].strip()
                    continue
        
        # Web sitesi kontrolü
        for pattern in website_patterns:
            website_match = re.search(pattern, line)
            if website_match:
                website = website_match.group(0)
                break
        
        # Eğer sadece "Web sitesi" yazıyorsa, bu bir link değil etikettir
        if line_lower == "web sitesi" or line_lower == "website":
            continue
        
        # Eğer web sitesi bulunmuşsa, sonraki adıma geç
        if website:
            continue
        
        # Adres kontrolü - tipik adres kalıpları
        is_address = False
        for pattern in address_patterns:
            if re.search(pattern, line):
                # Adres olarak işaretle ve başka bir şey olup olmadığını kontrol et
                if not address:
                    address = line
                    is_address = True
                    break
        
        if is_address:
            continue
        
        # Eğer adres belirlenmemişse ve bu satır işletme türünden hemen sonra geliyorsa, muhtemelen adrestir
        if not address and i == 2 and location_type:
            # Açılış/kapanış saati veya telefon numarası değilse adres olabilir
            if not any(x in line_lower for x in ["açık", "kapalı"]) and not re.search(r'\d+[,.]\d+\(\d+\)', line):
                address = line
                continue
        
        # Uzun metinler açıklama olabilir
        if len(line) > 100:
            description += line + "\n"
            continue
        
        # Kaydedilmemiş ve anlam ifade eden bir satır kaldıysa, açıklamaya ekle
        if len(line) > 3:
            description += line + "\n"
    
    # Tüm Bilgiler içinde telefon ve e-posta ara (eğer hâlâ bulunamadıysa)
    if not phone:
        for pattern in phone_patterns:
            matches = re.findall(pattern, item)
            if matches:
                for match in matches:
                    # Açılış-kapanış saati değilse
                    if not any(x in match for x in [":", "⋅"]):
                        phone = match
                        break
                if phone:
                    break
    
    # Telefonu tekrar kontrol et - Açılış saati yerine telefon göstermediğinden emin ol
    if phone and (re.match(r'^\d{1,2}:\d{2}$', phone) or "Kapanış saati" in phone or "Açılış zamanı" in phone):
        phone = ""
    
    # E-posta aramaya devam et
    if not email:
        email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', item)
        if email_matches:
            email = email_matches[0]
    
    # Web sitesi bulunamadıysa, ilgili metni ara
    if not website:
        for pattern in website_patterns:
            website_matches = re.findall(pattern, item)
            if website_matches:
                website = website_matches[0] if isinstance(website_matches[0], str) else website_matches[0][0]
                break
    
    # Ayrıştırılmış verileri ekle
    return {
        'Ülke': country,
        'Eyalet/Şehir': state,
        'Dil': language,
        'Şirket Adı': company_name,
        'İşletme Türü': location_type,
        'Adres': address,
        'Telefon': phone,
        'E-posta': email,
        'Web Sitesi': website,
        'Değerlendirme': rating,
        'Puan': rating_value,
        'Yorum Sayısı': review_count,
        'Açıklama': description,
        'Çalışma Saatleri': opening_hours,
        'Tüm Bilgiler': item
    }

# Main bloğuna iyileştirmeler ekliyorum
if __name__ == "__main__":
    # Komut satırı argümanına göre işlem yap
    import sys
    
    # Eğer "convert" argümanı varsa, mevcut JSON dosyalarını Excel'e dönüştür
    if len(sys.argv) > 1 and sys.argv[1] == "convert":
        convert_json_to_excel()
    # Eğer "city" argümanı varsa, tek bir şehir için arama yap
    elif len(sys.argv) > 1 and sys.argv[1] == "city":
        search_one_city()
    # Eğer "coal" argümanı varsa, kömür madenleri için özel arama yap
    elif len(sys.argv) > 1 and sys.argv[1] == "coal":
        print("Kömür madenleri ve maden ocakları arama işlemi başlıyor...")
        
        # İkinci argüman varsa, belirli bir ülke için arama yap
        if len(sys.argv) > 2:
            country = sys.argv[2]
            if country in COUNTRIES_INFO:
                print(f"Sadece {country} için arama yapılıyor...")
                # Sadece tek ülke için arama yapacak kod yapısı
                all_results = {}
                json_files = []
                
                info = COUNTRIES_INFO[country]
                country_lang = info['dil']
                print(f"\n=== {country} ÜLKESİ ARANIYOR (Dil: {country_lang}) ===")
                country_results = {}
                
                # Ülkenin kendi dilindeki kömür madeni ve maden ocağı terimlerini kullan
                if country_lang in COAL_MINE_TERMS:
                    search_terms = COAL_MINE_TERMS[country_lang]
                else:
                    # Eğer dil desteklenmiyorsa, İngilizce terimleri kullan
                    search_terms = COAL_MINE_TERMS['en']
                    print(f"Uyarı: {country_lang} dili desteklenmiyor, İngilizce terimler kullanılıyor.")
                
                # Önce genel ülke araması
                country_lang_results = []
                for term in search_terms:
                    search_query = f"{term} {country}"
                    print(f"Ülke Araması: {search_query}")
                    
                    try:
                        results = search(search_query)
                        if results:
                            country_lang_results.extend(results)
                            print(f"{len(results)} sonuç bulundu")
                    except Exception as e:
                        print(f"Hata oluştu: {e}")
                        continue
                
                if country_lang_results:
                    # Genel ülke sonuçlarını kaydet
                    json_file = save_results(country_lang_results, f"{country}_genel", country_lang)
                    json_files.append(json_file)
                    
                    # Ülkenin genel sonuçlarını ekle
                    country_results["genel"] = country_lang_results
                    
                # Şimdi eyalet/şehir bazlı aramalar
                state_results = {}
                for state in info['eyaletler']:
                    print(f"\n--- {state} eyaleti/şehri için arama yapılıyor ---")
                    state_lang_results = []
                    
                    for term in search_terms:
                        search_query = f"{term} {state} {country}"
                        print(f"Eyalet Araması: {search_query}")
                        
                        try:
                            results = search(search_query)
                            if results:
                                state_lang_results.extend(results)
                                print(f"{len(results)} sonuç bulundu")
                        except Exception as e:
                            print(f"Hata oluştu: {e}")
                            continue
                    
                    if state_lang_results:
                        state_results[state] = state_lang_results
                        # Eyalet sonuçlarını kaydet
                        json_file = save_results(state_lang_results, f"{country}_{state}", country_lang)
                        json_files.append(json_file)
                
                if state_results:
                    country_results["eyaletler"] = state_results
                
                if country_results:
                    all_results[country] = country_results
                
                # Tüm sonuçları kaydet
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                json_filename = f"{country}_coal_mines_{timestamp}.json"
                with open(json_filename, 'w', encoding='utf-8') as f:
                    json.dump(all_results, f, ensure_ascii=False, indent=2)
                print(f"\nTüm sonuçlar {json_filename} dosyasına kaydedildi.")
                json_files.append(json_filename)
                
                # JSON dosyalarını ayrı ayrı Excel'e dönüştür
                print("\nJSON dosyaları paralel olarak Excel'e dönüştürülüyor...")
                excel_files = []
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    future_to_file = {executor.submit(json_to_excel, file): file for file in json_files}
                    for future in concurrent.futures.as_completed(future_to_file):
                        result = future.result()
                        if result:
                            excel_files.append(result)
                            print(f"Dönüştürüldü: {result}")
                
                # Tek bir büyük Excel dosyası da oluştur
                print("\nAyrıca tüm kömür madeni sonuçları tek bir Excel dosyasına dönüştürülüyor...")
                excel_filename = create_combined_excel_coal_mines(all_results, timestamp)
                if excel_filename:
                    excel_files.append(excel_filename)
                    print(f"Tüm sonuçlar tek bir dosyada toplandı: {excel_filename}")
                
                # Oluşturulan tüm Excel dosyalarını listele
                print("\nOluşturulan Excel dosyaları:")
                for file in excel_files:
                    print(f"- {file}")
            else:
                print(f"Hata: {country} ülkesi tanımlı değil. Mevcut ülkeler:")
                for c in COUNTRIES_INFO.keys():
                    print(f"- {c}")
        else:
            # Tüm ülkeler için arama yap
            results = search_coal_mines_worldwide()
    else:
        print("Maden şirketleri arama işlemi başlıyor...")
        results = search_mining_companies()
